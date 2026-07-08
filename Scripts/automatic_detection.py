import argparse
import math
import os
import re

import cv2
import numpy as np
import pandas as pd
import pytesseract
import yaml
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pytesseract import Output
from ultralytics import YOLO


def calculate_area_per_frame(coordinates_directory, frame_width=1920, frame_height=1080, actual_distance_cm=10):
    """Convert per-frame laser-dot YOLO labels into visible seafloor area (m²) per frame."""
    data = []

    def calculate_distance(p1, p2):
        return math.sqrt((p1[0] - p2[0]) ** 2 + (p1[1] - p2[1]) ** 2)

    def extract_frame_number(filename):
        match = re.search(r'(\d+)', filename)
        return int(match.group(1)) if match else None

    for filename in os.listdir(coordinates_directory):
        if not filename.endswith(".txt"):
            continue
        frame_number = extract_frame_number(filename)
        if frame_number is None:
            continue
        with open(os.path.join(coordinates_directory, filename)) as f:
            lines = f.readlines()
        points = []
        for line in lines:
            parts = line.strip().split()
            if len(parts) >= 3:
                # YOLO format: class_id  x_center  y_center  width  height
                points.append((float(parts[1]) * frame_width, float(parts[2]) * frame_height))
        if len(points) == 2:
            distance_px = calculate_distance(points[0], points[1])
            cm_per_px = actual_distance_cm / distance_px
            area_m2 = (frame_width * cm_per_px / 100) * (frame_height * cm_per_px / 100)
            data.append([frame_number, area_m2])

    df = pd.DataFrame(data, columns=['Frame Number', 'Area (m^2)'])
    return df.sort_values(by='Frame Number')


def apply_alternating_row_colors(filepath):
    wb = load_workbook(filepath)
    ws = wb.active

    fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    current_fill = fill_white

    frame_col = None
    for idx, cell in enumerate(ws[1]):
        if cell.value == 'Frame Number':
            frame_col = idx
            break

    last_frame = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        val = row[frame_col].value if frame_col is not None and frame_col < len(row) else None
        if val != last_frame:
            current_fill = fill_green if current_fill == fill_white else fill_white
            last_frame = val
        for cell in row:
            cell.fill = current_fill

    wb.save(filepath)
    print(f'Row colors applied: {filepath}')


def analyze_video(video_meta, cfg, model):
    """Process a single ROV video: extract frames, run YOLO detection, OCR metadata, compute density."""
    video_path = video_meta['video_path']
    laser_dir = video_meta['laser_coordinates_dir']
    rov_dive = video_meta['rov_dive']

    actual_laser_distance_cm = video_meta.get('actual_laser_distance_cm', cfg.get('actual_laser_distance_cm', 10))
    frame_interval = cfg.get('frame_interval_seconds', 30)
    conf_threshold = cfg.get('conf_threshold', 0.1)
    iou_threshold = cfg.get('iou_threshold', 0.1)
    roi_heading_rel = tuple(cfg['roi_heading_rel'])
    roi_depth_rel = tuple(cfg['roi_depth_rel'])
    output_dir_base = os.path.join(cfg['output_dir'], rov_dive)

    metadata_dict = {
        "Date": video_meta.get('date', ''),
        "Start Time": video_meta.get('start_time', ''),
        "End Time": video_meta.get('end_time', ''),
        "ROV Dive": rov_dive,
        "Total video time": video_meta.get('total_time', ''),
        "Total Video Time (s)": video_meta.get('total_video_time_seconds', ''),
        "Distance Covered (m)": video_meta.get('distance_covered', ''),
        "Start Latitude": video_meta.get('start_latitude', ''),
        "Start Longitude": video_meta.get('start_longitude', ''),
        "Type": video_meta.get('transect', ''),
        "Temperature (C)": video_meta.get('temperature', ''),
        "Habitat_type": video_meta.get('habitat_type', ''),
    }

    def get_unique_directory_name(base):
        counter, unique = 1, base
        while os.path.exists(unique):
            unique = f"{base}_{counter}"
            counter += 1
        return unique

    def get_unique_filename(directory, filename):
        base, ext = os.path.splitext(filename)
        counter, unique = 1, filename
        while os.path.exists(os.path.join(directory, unique)):
            unique = f"{base}_{counter}{ext}"
            counter += 1
        return unique

    output_dir = get_unique_directory_name(output_dir_base)
    frames_dir = os.path.join(output_dir, 'frames')
    excel_path = os.path.join(output_dir, f'{rov_dive}_results.xlsx')
    os.makedirs(frames_dir, exist_ok=True)

    def random_color():
        return tuple(int(x) for x in np.random.choice(range(256), size=3))

    def relative_center_to_bbox(frame, center_rel):
        h, w = frame.shape[:2]
        cx, cy, wrel, hrel = center_rel
        x1 = max(0, int(cx * w) - int(wrel * w) // 2)
        y1 = max(0, int(cy * h) - int(hrel * h) // 2)
        x2 = min(w, int(cx * w) + int(wrel * w) // 2)
        y2 = min(h, int(cy * h) + int(hrel * h) // 2)
        return x1, y1, x2, y2

    def preprocess_roi(roi, method='default'):
        if method == 'resize':
            return cv2.resize(roi, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        elif method == 'grayscale':
            return cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        elif method == 'clahe':
            gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
            return clahe.apply(gray)
        elif method == 'binary':
            gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
            _, result = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            return result
        else:
            resized = cv2.resize(roi, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
            gray = cv2.GaussianBlur(cv2.cvtColor(resized, cv2.COLOR_BGR2GRAY), (3, 3), 0)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
            _, result = cv2.threshold(clahe.apply(gray), 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            return result

    def extract_text_from_region(frame, region):
        x1, y1, x2, y2 = region
        roi = frame[y1:y2, x1:x2]
        results = []
        for method in ['default', 'resize', 'grayscale', 'clahe', 'binary']:
            preprocessed = preprocess_roi(roi, method=method)
            text = pytesseract.image_to_string(
                preprocessed, config=r'--oem 3 --psm 6', output_type=Output.STRING
            )
            results.append(text.strip())
        return results

    def format_text(text, pattern):
        match = re.search(pattern, text)
        return match.group(0) if match else ''

    depth_pattern = r'\d{2,3}\.\d{2}'
    heading_pattern = r'\d{3}'

    if not os.path.exists(video_path):
        print(f'Video not found: {video_path}')
        return None

    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        print(f'Error opening video: {video_path}')
        return None

    fps = cap.get(cv2.CAP_PROP_FPS)
    frame_step = max(1, int(frame_interval * fps))
    frame_count = 0
    results_list = []

    area_df = calculate_area_per_frame(laser_dir, actual_distance_cm=actual_laser_distance_cm)
    area_dict = area_df.set_index('Frame Number')['Area (m^2)'].to_dict()

    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break

        if frame_count % frame_step == 0:
            frame_number = (frame_count // frame_step) + 1
            print(f'Processing frame {frame_number}')

            roi_heading = relative_center_to_bbox(frame, roi_heading_rel)
            roi_depth = relative_center_to_bbox(frame, roi_depth_rel)

            heading_results = extract_text_from_region(frame, roi_heading)
            depth_results = extract_text_from_region(frame, roi_depth)

            heading = next((format_text(r, heading_pattern) for r in heading_results
                            if format_text(r, heading_pattern)), 'N/A')
            depth = next((format_text(r, depth_pattern) for r in depth_results
                          if format_text(r, depth_pattern)), 'N/A')
            print(f'Heading: {heading}, Depth: {depth}')

            det = model(frame, conf=conf_threshold, iou=iou_threshold)[0]
            coral_count = len(det.masks.data) if det.masks is not None else 0

            if coral_count > 0:
                for idx, (mask_data, box) in enumerate(zip(det.masks.data, det.boxes)):
                    score = round(box.conf.cpu().numpy()[0], 2)
                    color = random_color()
                    mask_resized = cv2.resize(
                        mask_data.cpu().numpy().astype(np.uint8),
                        (frame.shape[1], frame.shape[0]),
                    )
                    contours, _ = cv2.findContours(mask_resized, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    if contours:
                        contour = max(contours, key=cv2.contourArea)
                        cv2.drawContours(frame, [contour], -1, color, 2)
                        x, y, w, h = cv2.boundingRect(contour)
                        cv2.putText(frame, f'Coral {idx + 1} ({score:.2f})',
                                    (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

            results_list.append({
                'Frame Number': frame_number,
                'Coral Count': coral_count,
                'Heading': heading,
                'Depth': depth,
                'Area (m^2)': area_dict.get(frame_number, 'N/A'),
            })

            frame_path = os.path.join(frames_dir, get_unique_filename(frames_dir, f'frame_{frame_number}.jpg'))
            cv2.imwrite(frame_path, frame)
            print(f'Saved: {frame_path}')

        frame_count += 1

    cap.release()

    if not results_list:
        print(f'No frames processed: {video_path}')
        return None

    df = pd.DataFrame(results_list)
    meta_df = pd.concat([pd.DataFrame([metadata_dict])] * len(df), ignore_index=True)
    df_full = pd.concat([meta_df, df], axis=1)

    df_full['Coral Count'] = pd.to_numeric(df_full['Coral Count'], errors='coerce').fillna(0)
    df_full['Area (m^2)'] = pd.to_numeric(df_full['Area (m^2)'], errors='coerce').fillna(0)
    df_full['Coral Density'] = df_full['Coral Count'] / df_full['Area (m^2)'].replace(0, float('nan'))

    total_corals = df_full['Coral Count'].sum()
    total_area = df_full['Area (m^2)'].sum()
    total_density = total_corals / total_area if total_area > 0 else 0

    last = len(df_full) - 1
    df_full.at[last, 'Total Coral Count'] = total_corals
    df_full.at[last, 'Total Area (m^2)'] = total_area
    df_full.at[last, 'Total Coral Density'] = total_density

    df_full.to_excel(excel_path, index=False)
    apply_alternating_row_colors(excel_path)
    print(f'Results saved: {excel_path}')
    return excel_path


def main():
    parser = argparse.ArgumentParser(
        description="ROV video analysis: YOLO detection + OCR + laser-calibrated coral density.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--config", required=True, help="Path to video metadata YAML")
    args = parser.parse_args()

    with open(args.config, encoding="utf-8-sig") as f:
        cfg = yaml.safe_load(f)

    tesseract_path = cfg.get('tesseract_path')
    if tesseract_path:
        pytesseract.pytesseract.tesseract_cmd = tesseract_path

    os.makedirs(cfg['output_dir'], exist_ok=True)

    model_path = cfg['model_path']
    if not os.path.exists(model_path):
        raise FileNotFoundError(f"YOLO model not found: {model_path}")
    model = YOLO(model_path)

    result_files = []
    for video_meta in cfg['videos']:
        result = analyze_video(video_meta, cfg, model)
        if result:
            result_files.append(result)

    if not result_files:
        print("No results to combine.")
        return

    combined_df = pd.concat([pd.read_excel(f) for f in result_files], ignore_index=True)
    combined_path = cfg.get('combined_output_path', os.path.join(cfg['output_dir'], 'combined_results.xlsx'))
    os.makedirs(os.path.dirname(os.path.abspath(combined_path)), exist_ok=True)
    combined_df.to_excel(combined_path, index=False)
    apply_alternating_row_colors(combined_path)
    print(f'Combined results saved: {combined_path}')

    # Summary analysis across all videos
    coral_data = pd.read_excel(combined_path)
    coral_data = coral_data.rename(columns={'Coral Count': 'Coral_Count'})

    summary_path = os.path.join(cfg['output_dir'], 'summary_analysis.xlsx')
    with pd.ExcelWriter(summary_path, engine='openpyxl') as writer:
        for group_col, sheet in [
            ('Habitat_type', 'By_Habitat'),
            ('Type', 'By_Transect_Type'),
        ]:
            if group_col in coral_data.columns:
                coral_data.groupby(group_col).agg({'Coral_Count': 'sum'}).reset_index() \
                    .to_excel(writer, sheet_name=sheet, index=False)

        for group_col, sheet in [
            ('ROV Dive', 'Density_by_Dive'),
            ('Habitat_type', 'Density_by_Habitat'),
            ('Type', 'Density_by_Type'),
        ]:
            if group_col in coral_data.columns and 'Area (m^2)' in coral_data.columns:
                grp = coral_data.groupby(group_col).agg(
                    {'Coral_Count': 'sum', 'Area (m^2)': 'sum'}
                ).reset_index()
                grp['Coral_Density'] = grp['Coral_Count'] / grp['Area (m^2)']
                grp.to_excel(writer, sheet_name=sheet, index=False)

    print(f'Summary analysis saved: {summary_path}')


if __name__ == "__main__":
    main()
