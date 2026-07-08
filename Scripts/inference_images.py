import argparse
import math
import os
import re

import cv2
import numpy as np
import pandas as pd
import pytesseract
import torch
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pytesseract import Output
from ultralytics import YOLO

ALLOWED_IMAGE_EXTENSIONS = ('.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.tif')


def apply_alternating_row_colors(filepath):
    try:
        wb = load_workbook(filepath)
        ws = wb.active

        fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        current_fill = fill_white

        image_index_col = None
        for idx, cell in enumerate(ws[1]):
            if cell.value and isinstance(cell.value, str) and cell.value.strip().lower() == 'image index':
                image_index_col = idx
                break

        if image_index_col is None:
            print(f"Error: 'Image Index' column not found in {filepath}.")
            return

        last_index = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            if image_index_col < len(row):
                current_index = row[image_index_col].value
                if current_index != last_index:
                    current_fill = fill_green if current_fill == fill_white else fill_white
                    last_index = current_index
                for cell in row:
                    cell.fill = current_fill

        wb.save(filepath)
        print(f'Alternating row colors applied: {filepath}')
    except Exception as e:
        print(f"Error applying colors: {e}")


def analyze_images_in_folder(image_folder_path, model_path, conf_threshold, iou_threshold,
                              area_threshold_percent, tesseract_path, roi_depth_rel,
                              output_dir, excel_filename):

    def get_unique_directory_name(base):
        counter = 1
        unique = base
        while os.path.exists(unique):
            unique = f"{base}_{counter}"
            counter += 1
        return unique

    def get_unique_filename(directory, filename):
        base, ext = os.path.splitext(filename)
        counter = 1
        unique = filename
        while os.path.exists(os.path.join(directory, unique)):
            unique = f"{base}_{counter}{ext}"
            counter += 1
        return unique

    input_folder_name = os.path.basename(os.path.normpath(image_folder_path))
    output_dir_base = os.path.join(output_dir, f"{input_folder_name}_Analysis_Output")
    output_dir_final = get_unique_directory_name(output_dir_base)
    frames_dir = os.path.join(output_dir_final, 'annotated_images')
    excel_path = os.path.join(output_dir_final, excel_filename)

    try:
        os.makedirs(output_dir_final, exist_ok=True)
        os.makedirs(frames_dir, exist_ok=True)
        print(f"Output directory: {output_dir_final}")
    except OSError as e:
        print(f"Error creating output directories: {e}")
        return

    # Tesseract setup
    try:
        if tesseract_path:
            pytesseract.pytesseract.tesseract_cmd = tesseract_path
        tesseract_version = pytesseract.get_tesseract_version()
        print(f"Tesseract version {tesseract_version} found.")
    except FileNotFoundError:
        print(f"Error: Tesseract not found. Install it or pass --tesseract_path.")
        return
    except Exception as e:
        print(f"Error initializing Tesseract: {e}")
        return

    # YOLO model setup
    try:
        if not os.path.exists(model_path):
            raise FileNotFoundError(f"YOLO model not found: {model_path}")
        model = YOLO(model_path)
        print(f"YOLO model loaded: {model_path}")
    except Exception as e:
        print(f"Error loading YOLO model: {e}")
        return

    def random_color():
        return tuple(int(x) for x in np.random.choice(range(256), size=3))

    def relative_center_to_bbox(frame_shape, center_rel):
        h, w = frame_shape[:2]
        cx, cy, wrel, hrel = center_rel
        x1 = max(0, int(cx * w) - int(wrel * w) // 2)
        y1 = max(0, int(cy * h) - int(hrel * h) // 2)
        x2 = min(w, int(cx * w) + int(wrel * w) // 2)
        y2 = min(h, int(cy * h) + int(hrel * h) // 2)
        return x1, y1, x2, y2

    def preprocess_roi(roi, method='default'):
        if roi is None or roi.size == 0:
            return None
        try:
            if method == 'resize':
                return cv2.resize(roi, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC) if roi.shape[0] > 0 and roi.shape[1] > 0 else roi
            elif method == 'grayscale':
                return cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
            elif method == 'clahe':
                gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                return clahe.apply(gray)
            elif method == 'binary':
                gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
                return cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
            else:
                if roi.shape[0] > 0 and roi.shape[1] > 0:
                    resized = cv2.resize(roi, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
                    gray = cv2.GaussianBlur(cv2.cvtColor(resized, cv2.COLOR_BGR2GRAY), (3, 3), 0)
                    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                    _, result = cv2.threshold(clahe.apply(gray), 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
                    return result
                return roi
        except Exception as e:
            print(f"Preprocessing error ({method}): {e}")
            return roi

    def extract_text_from_region(frame, region_coords):
        x1, y1, x2, y2 = region_coords
        if x1 >= x2 or y1 >= y2:
            return ['N/A'] * 5
        roi = frame[y1:y2, x1:x2]
        if roi.size == 0:
            return ['N/A'] * 5
        results = []
        for method in ['default', 'resize', 'grayscale', 'clahe', 'binary']:
            preprocessed = preprocess_roi(roi.copy(), method=method)
            if preprocessed is None or preprocessed.size == 0:
                results.append("PreProc_Error")
                continue
            try:
                text = pytesseract.image_to_string(
                    preprocessed,
                    config=r'--oem 3 --psm 7 -c tessedit_char_whitelist=0123456789.',
                    output_type=Output.STRING,
                    timeout=5,
                )
                results.append(re.sub(r'[^\d.]', '', text).strip())
            except Exception:
                results.append("OCR_Error")
        return results

    def format_text(text, pattern):
        if not isinstance(text, str):
            return ''
        match = re.search(pattern, text)
        return match.group(0) if match else ''

    depth_pattern = r'\d{1,3}\.\d{1,2}'

    if not os.path.isdir(image_folder_path):
        print(f"Not a directory: {image_folder_path}")
        return

    image_files = [f for f in os.listdir(image_folder_path) if f.lower().endswith(ALLOWED_IMAGE_EXTENSIONS)]
    if not image_files:
        print(f"No images found in: {image_folder_path}")
        return

    results_list = []
    print(f"Processing {len(image_files)} images from: {image_folder_path}")
    print(f"Area filter: detections > {area_threshold_percent * 100:.1f}% of frame are rejected.")

    for idx, image_filename in enumerate(image_files, start=1):
        image_path = os.path.join(image_folder_path, image_filename)
        try:
            frame = cv2.imread(image_path)
            if frame is None:
                print(f"Warning: could not read {image_filename}. Skipping.")
                continue

            print(f'\n--- Image {idx}/{len(image_files)}: {image_filename} ---')
            total_area = frame.shape[0] * frame.shape[1]
            max_area = area_threshold_percent * total_area

            roi_depth_coords = relative_center_to_bbox(frame.shape, roi_depth_rel)
            depth_ocr = extract_text_from_region(frame.copy(), roi_depth_coords)
            depth = next((formatted for r in depth_ocr if (formatted := format_text(r, depth_pattern))), 'N/A')
            print(f'Depth OCR: {depth_ocr} -> {depth}')

            annotated = frame.copy()
            coral_count = 0
            try:
                det = model(annotated, conf=conf_threshold, iou=iou_threshold, verbose=False, agnostic_nms=True)[0]
                has_masks = det.masks is not None and len(det.masks.data) > 0
                has_boxes = det.boxes is not None and len(det.boxes.data) > 0

                if has_masks and has_boxes and len(det.masks.data) == len(det.boxes.data):
                    for det_idx, (mask_data, box) in enumerate(zip(det.masks.data, det.boxes)):
                        score = round(box.conf.cpu().numpy()[0], 2)
                        color = random_color()
                        mask_resized = cv2.resize(
                            mask_data.cpu().numpy().astype(np.uint8),
                            (frame.shape[1], frame.shape[0]),
                            interpolation=cv2.INTER_NEAREST,
                        )
                        contours, _ = cv2.findContours(mask_resized, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                        if contours:
                            contour = max(contours, key=cv2.contourArea)
                            if cv2.contourArea(contour) <= max_area:
                                coral_count += 1
                                cv2.drawContours(annotated, [contour], -1, color, 2)
                                x, y, w, h = cv2.boundingRect(contour)
                                cv2.putText(annotated, f'Coral {coral_count} ({score:.2f})',
                                            (x, max(y - 10, 10)), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
                            else:
                                print(f"  Detection {det_idx + 1} rejected (area too large)")
                elif has_boxes:
                    coral_count = len(det.boxes.data)
                    print("Warning: no masks found, using box count.")
                else:
                    print("No detections.")
            except Exception as e:
                print(f"Detection error for {image_filename}: {e}")
                coral_count = 'Error'

            results_list.append({
                'Image Index': idx,
                'Image Name': image_filename,
                'Coral Count': coral_count,
                'Depth': depth,
            })

            save_name = get_unique_filename(frames_dir, f'annotated_{os.path.splitext(image_filename)[0]}.jpg')
            cv2.imwrite(os.path.join(frames_dir, save_name), annotated)

        except Exception as e:
            print(f"Error processing {image_filename}: {e}")
            results_list.append({'Image Index': idx, 'Image Name': image_filename,
                                  'Coral Count': 'Error', 'Depth': 'Error'})

    print("\nFinished processing images.")
    if results_list:
        try:
            df = pd.DataFrame(results_list, columns=['Image Index', 'Image Name', 'Coral Count', 'Depth'])
            df['Coral Count'] = pd.to_numeric(df['Coral Count'], errors='coerce').fillna(0).astype(int)
            df.to_excel(excel_path, index=False, engine='openpyxl')
            print(f'Results saved: {excel_path}')
            apply_alternating_row_colors(excel_path)
        except Exception as e:
            print(f"Error saving Excel: {e}")
    else:
        print("No results to save.")


def main():
    parser = argparse.ArgumentParser(
        description="Run YOLO instance segmentation inference on a folder of still images.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--images", required=True, help="Path to folder containing images")
    parser.add_argument("--model", required=True, help="Path to YOLO model weights (.pt)")
    parser.add_argument("--conf", type=float, default=0.1, help="Confidence threshold")
    parser.add_argument("--iou", type=float, default=0.0, help="IOU threshold")
    parser.add_argument("--area_threshold", type=float, default=0.15,
                        help="Reject detections covering more than this fraction of the image")
    parser.add_argument("--tesseract_path", default=None,
                        help="Path to tesseract.exe (optional; uses system PATH if omitted)")
    parser.add_argument("--roi_depth", nargs=4, type=float,
                        default=[0.9501, 0.3522, 0.0341, 0.0294],
                        metavar=("CX", "CY", "W", "H"),
                        help="Depth OCR region in relative coords (center_x center_y width height)")
    parser.add_argument("--output_dir", default=None,
                        help="Directory for output (defaults to parent of --images folder)")
    parser.add_argument("--excel_name", default="results.xlsx", help="Output Excel filename")
    args = parser.parse_args()

    if not os.path.isdir(args.images):
        print(f"ERROR: images folder not found: {args.images}")
        return
    if not os.path.exists(args.model):
        print(f"ERROR: model file not found: {args.model}")
        return

    output_dir = args.output_dir or os.path.dirname(os.path.abspath(args.images))
    analyze_images_in_folder(
        args.images,
        args.model,
        args.conf,
        args.iou,
        args.area_threshold,
        args.tesseract_path,
        tuple(args.roi_depth),
        output_dir,
        args.excel_name,
    )


if __name__ == "__main__":
    main()
